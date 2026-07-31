#!/usr/bin/env python3
"""動作確認用のサンプル Excel を 3 つ作る。

    python3 tools/make_samples.py [出力先ディレクトリ]

既定の出力先は samples/。以下のケースを一通り含める。
  - 一致（同一キーが複数明細に分かれる = サマル必須）
  - 数量差異
  - 実棚のみ / SAPのみ
  - 変換表にヒットする倉庫 / ヒットせず元名称を使う倉庫
  - "1,234" 形式の文字列数値、小数の基本数量
  - 使用しない列にダミー値を入れて列ずれを検出できるようにする
"""

import sys
from pathlib import Path

from openpyxl import Workbook

ROOT = Path(__file__).resolve().parent.parent


def build_actual(path: Path) -> None:
    """実棚: A列=倉庫名 / E列=型式 / I列=台数"""
    wb = Workbook()
    ws = wb.active
    ws.title = "実棚"

    ws.append(["倉庫名", "棚番", "担当者", "確認日", "型式", "メーカー", "区分", "備考", "台数"])

    rows = [
        # 倉庫名, 型式, 台数（間の列はダミー）
        ("東京第一倉庫", "AAA-100", 3),
        ("東京第一倉庫", "AAA-100", 2),      # 同一キー → 合計 5 で SAP と一致
        ("東京第一倉庫", "BBB-200", 10),     # SAP は 8 → 差異 +2
        ("東京第一倉庫", "CCC-300", 4),      # SAP に無し → 実棚のみ
        ("大阪倉庫", "AAA-100", "1,200"),    # 文字列の桁区切り
        ("大阪倉庫", "DDD-400", 6),
        ("福岡倉庫", "EEE-500", 1.5),        # 小数
        ("ST99", "FFF-600", 7),              # 変換表に無い保管場所と突合
        ("名古屋倉庫", "GGG-700", 12),       # SAP 側が 2 明細に分かれる
    ]

    for i, (warehouse, model, qty) in enumerate(rows, start=1):
        ws.append([
            warehouse, f"A-{i:03d}", "山田", "2026-07-31", model,
            "サンプル製作所", "通常", "", qty,
        ])

    ws.freeze_panes = "A2"
    wb.save(path)


def build_sap(path: Path) -> None:
    """SAP: E列=保管場所名 / G列=品目テキスト / J列=基本数量"""
    wb = Workbook()
    ws = wb.active
    ws.title = "在庫一覧"

    ws.append([
        "プラント", "プラント名", "保管場所", "品目コード", "保管場所名",
        "品目グループ", "品目テキスト", "在庫タイプ", "単位", "基本数量",
    ])

    rows = [
        # 保管場所名(変換前), 品目テキスト, 基本数量
        ("ST01", "AAA-100", 5),
        ("ST01", "BBB-200", 8),             # 実棚 10 → 差異
        ("ST02", "AAA-100", "1,200"),
        ("ST02", "DDD-400", 6),
        ("ST02", "HHH-800", 9),             # 実棚に無し → SAPのみ
        ("ST03", "EEE-500", 1.5),
        ("ST99", "FFF-600", 4),
        ("ST99", "FFF-600", 3),             # 合計 7 で実棚と一致（変換表に無い）
        ("ST04", "GGG-700", 5),
        ("ST04", "GGG-700", 7),             # 合計 12 で実棚と一致
    ]

    for i, (location, item, qty) in enumerate(rows, start=1):
        ws.append([
            "1000", "サンプル工場", location[-2:], f"MAT{i:05d}", location,
            "G01", item, "無制限使用", "TAI", qty,
        ])

    ws.freeze_panes = "A2"
    wb.save(path)


def build_mapping(path: Path) -> None:
    """変換表: A列=変換前 / B列=変換後"""
    wb = Workbook()
    ws = wb.active
    ws.title = "変換表"

    ws.append(["変換前", "変換後"])
    for before, after in [
        ("ST01", "東京第一倉庫"),
        ("ST02", "大阪倉庫"),
        ("ST03", "福岡倉庫"),
        ("ST04", "名古屋倉庫"),
        # ST99 はあえて未登録 → 元の名前のまま照合される
    ]:
        ws.append([before, after])

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    out_dir = Path(sys.argv[1]) if len(sys.argv) > 1 else ROOT / "samples"
    out_dir.mkdir(parents=True, exist_ok=True)

    build_actual(out_dir / "sample_実棚.xlsx")
    build_sap(out_dir / "sample_SAP.xlsx")
    build_mapping(out_dir / "sample_変換表.xlsx")

    print(f"サンプルを作成しました: {out_dir}")
    for p in sorted(out_dir.iterdir()):
        print(f"  - {p.name}")


if __name__ == "__main__":
    main()
